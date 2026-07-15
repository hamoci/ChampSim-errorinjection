#!/usr/bin/env python3
"""Dump raw simulation metrics from results/normal_evaluation/ into one XLSX.

Sheet 1 (Definitions): glossary for every column appearing in later sheets,
  plus a short experiment overview.
Sheets 2..N: one per experiment subdirectory of results/normal_evaluation/.
  Each row is one sim run; columns are the run's configuration + every metric
  parsed from the .txt output that any figure ever consumes.

Run:
    python3 normal_evaluation_script/export_normal_evaluation.py
Output:
    normal_evaluation_script/export/normal_evaluation_raw.xlsx
"""

from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
RESULTS_ROOT = os.path.join(BASE_DIR, "results", "normal_evaluation")
OUT_DIR = os.path.join(SCRIPT_DIR, "export")
OUT_XLSX = os.path.join(OUT_DIR, "normal_evaluation_raw.xlsx")


# ---------------------------------------------------------------------------
# Filename parsers — pull the configuration columns out of each result name.
# ---------------------------------------------------------------------------

TRACE = r"(?P<trace>\d+\.[A-Za-z0-9_]+-\d+B)\.txt$"

_RETIRE_RE = re.compile(rf"^retire_(?P<pinning>on|off)_(?P<retirement_threshold>\d+)_(?P<error_rate>1e-\d+)_{TRACE}")

FILENAME_PARSERS: Dict[str, re.Pattern] = {
    "1_error_rate_sweep":              re.compile(rf"^pin_(?P<pinning>on|off)_(?P<error_rate>1e-\d+)_{TRACE}"),
    "2_retirement_threshold":          _RETIRE_RE,
    "2_retirement_threshold_only2":    _RETIRE_RE,
    "2_retirement_threshold_only32":   _RETIRE_RE,
    "2_retirement_threshold_4_8_16":   _RETIRE_RE,
    "3_error_way_capacity":            re.compile(rf"^errway_(?P<allocated_error_ways>\d+)w_(?P<error_rate>1e-\d+)_{TRACE}"),
    "4_llc_size_baseline":             re.compile(rf"^llc_baseline_(?P<llc_size_mb>\d+)MB_{TRACE}"),
    "5_llc_size_sensitivity":          re.compile(rf"^llc_(?P<llc_size_mb>\d+)MB_(?P<error_rate>1e-\d+)_{TRACE}"),
    "6_llc_way_sweep":                 re.compile(rf"^sweep_(?P<llc_size_mb>\d+)MB_w(?P<allocated_error_ways>\d+)_(?P<error_rate>1e-\d+)_{TRACE}"),
    "7_no_error_way_sweep":            re.compile(rf"^noerr_(?P<llc_size_mb>\d+)MB_w(?P<llc_ways>\d+)_{TRACE}"),
}

WORKLOAD_RE = re.compile(r"^(\d+\.\w+)")  # matches "602.gcc_s" — stops at "-" before the simpoint count


# ---------------------------------------------------------------------------
# Metric regexes — same patterns the figure scripts use, kept in one place.
# ---------------------------------------------------------------------------

RE_ROI = re.compile(r"CPU 0 cumulative IPC:\s+([\d.]+)\s+instructions:\s+(\d+)\s+cycles:\s+(\d+)")
RE_BR = re.compile(r"CPU 0 Branch Prediction Accuracy:\s+([\d.]+)%\s+MPKI:\s+([\d.]+)")
RE_LLC_LOAD = re.compile(r"cpu0->LLC LOAD\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+)")
RE_LLC_TOTAL = re.compile(r"cpu0->LLC TOTAL\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+)")
RE_ROW_BUFFER_MISS = re.compile(r"ROW_BUFFER_MISS:\s+(\d+)")
RE_PAGE_SIZE = re.compile(r"Page size:\s+(\d+)")

# Error model
RE_ERR_INTERVAL = re.compile(r"Error cycle interval:\s+(\d+)\s+cycles")
RE_ERR_PER_INT = re.compile(r"Errors per interval:\s+(\d+)")
RE_TOTAL_ERR = re.compile(r"Total Error Accesses:\s+(\d+)")
RE_TOTAL_DRAM_ERROR_EVENTS = re.compile(r"\[ERROR\]\s+Total DRAM Error Events:\s+(\d+)")
RE_TOTAL_KNOWN = re.compile(r"Total Known Error Addresses:\s+(\d+)")
RE_BASELINE_RETIRE = re.compile(r"Baseline Retirement Threshold:\s+(\d+)")
RE_BASELINE_PAGES = re.compile(r"Baseline Page Retirements:\s+(\d+)")
RE_ERROR_THRESH = re.compile(r"\[ERROR\]\s+Retirement Threshold:\s+(\d+)")
RE_ERROR_PAGES_NTH = re.compile(r"\[ERROR\]\s+Page Retirements \(\d+(?:th|st|nd|rd) err\):\s+(\d+)")
RE_ERROR_PAGES_RETIRED = re.compile(r"\[ERROR\]\s+Pages Retired:\s+(\d+)")

# LLC error-way bookkeeping (only emitted when pinning is ON)
RE_ALLOC = re.compile(r"\[LLC\]\s+Allocated Error Ways per Set:\s+(\d+)")
RE_MAX = re.compile(r"\[LLC\]\s+Max Error Ways per Set:\s+(\d+)")
RE_TOTAL_SLOTS = re.compile(r"\[LLC\]\s+Total Error Way Slots:\s+(\d+)")
RE_USED_SLOTS = re.compile(r"\[LLC\]\s+Used Slots:\s+(\d+)\s+\(([\d.]+)%\)")
RE_UNUSED_SLOTS = re.compile(r"\[LLC\]\s+Unused Slots:\s+(\d+)\s+\(([\d.]+)%\)")
RE_EW_HITS = re.compile(r"\[LLC\]\s+Error Way Hits:\s+(\d+)")
RE_EW_FILLS = re.compile(r"\[LLC\]\s+Error Way Fills.*?:\s+(\d+)")
RE_EW_HITRATE = re.compile(r"\[LLC\]\s+Error Way Hit Rate:\s+([\d.]+)%")
RE_EW_EVICT = re.compile(r"\[LLC\]\s+Error Way Evictions.*?:\s+(\d+)")
RE_PINNED = re.compile(r"\[LLC\]\s+Pinned in Error Way:\s+(\d+)\s+\(([\d.]+)%\)")
RE_IN_NORMAL_WAY = re.compile(r"\[LLC\]\s+In Normal Way \(unprotected\):\s+(\d+)")
RE_NOT_IN_LLC = re.compile(r"\[LLC\]\s+Not in LLC \(DRAM exposed\):\s+(\d+)")

# Pinning OFF: Baseline Protection Coverage block (only in 2_retirement_threshold_*
# subsidiary dirs — the simulator added this block in newer reruns).
RE_BASELINE_RETIRED = re.compile(r"\[LLC\]\s+Retired \(page offline\):\s+(\d+)\s+\(([\d.]+)%\)")
RE_BASELINE_LIVE = re.compile(r"\[LLC\]\s+Live \(still tracked\):\s+(\d+)")


# ---------------------------------------------------------------------------
# Per-experiment column ordering. Configuration columns first, then metrics.
# Build from a few shared bundles to keep meanings consistent across sheets.
# ---------------------------------------------------------------------------

BASE_COLS = [
    "workload", "trace", "filename", "panic",
    "ipc", "instructions", "cycles",
    "branch_accuracy_pct", "branch_mpki",
    "llc_load_access", "llc_load_hit", "llc_load_miss",
    "llc_total_access", "llc_total_hit", "llc_total_miss",
    "llc_load_mpki", "rbmpki",
    "page_size_bytes",
]
ERROR_COLS = [
    "error_cycle_interval", "errors_per_interval",
    "total_error_accesses", "total_dram_error_events",
    "total_known_error_addresses",
    "page_retirements_nth_err", "pages_retired",
    "baseline_retirement_threshold", "baseline_page_retirements",
    "error_retirement_threshold",
]
PIN_COLS = [
    "allocated_error_ways_per_set", "max_error_ways_per_set",
    "total_error_way_slots",
    "used_slots", "used_slots_pct",
    "unused_slots", "unused_slots_pct",
    "error_way_hits", "error_way_fills", "error_way_hit_rate_pct",
    "error_way_evictions",
    "pinned_in_error_way", "pinned_in_error_way_pct",
    "in_normal_way_unprotected", "not_in_llc_dram_exposed",
]
# Baseline Protection Coverage block (pin OFF runs in 2_retirement_threshold_* subsidiary dirs).
PIN_OFF_COVERAGE_COLS = [
    "baseline_retired_count_page_offline",
    "baseline_retired_pct_page_offline",
    "baseline_live_still_tracked",
]
# Derived (computed in code, not parsed):
#   retired_lines_inferred = max(0, total_dram_error_events - total_known_error_addresses)
#   protected_lines_pct    = fig9 formula, picks pin_on or pin_off variant automatically
DERIVED_PROTECTION_COLS = [
    "retired_lines_inferred",
    "protected_lines_pct",
]


@dataclass
class ExperimentSpec:
    dirname: str               # subdir under results/normal_evaluation/
    sheet: str                 # sheet name (<=31 chars)
    config_cols: List[str]     # config columns from the filename (in display order)
    metric_cols: List[str]     # metric columns (in display order)
    sort_keys: List[str]       # columns to sort by inside the sheet
    summary: str               # one-line description used on the Definitions sheet


EXPERIMENTS: List[ExperimentSpec] = [
    ExperimentSpec(
        dirname="1_error_rate_sweep",
        sheet="1_error_rate_sweep",
        config_cols=["pinning", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + PIN_OFF_COVERAGE_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["pinning", "error_rate", "workload"],
        summary="CE-rate sweep (1e-5 → 1e-8) with pinning ON and OFF. pin_off rows use threshold=1; for pin_off coverage data use the 2_retirement_threshold_* sheets instead. Drives fig6/fig8/fig8b/fig13.",
    ),
    ExperimentSpec(
        dirname="2_retirement_threshold",
        sheet="2_retirement_threshold",
        config_cols=["pinning", "retirement_threshold", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + PIN_OFF_COVERAGE_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["pinning", "retirement_threshold", "error_rate", "workload"],
        summary="Page-offline retirement threshold sweep with pinning ON/OFF across all CE rates. Includes [LLC] Baseline Protection Coverage block for pin_off → protected_lines_pct is populated for completed runs. Drives fig7/7b/7c/7d (and fig6/fig8/fig13 via retire_off thr=2 = Conventional Page Offline; fig9 pin-off marker via retire_off thr=32).",
    ),
    ExperimentSpec(
        dirname="2_retirement_threshold_only2",
        sheet="2_retirement_threshold_only2",
        config_cols=["pinning", "retirement_threshold", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + PIN_OFF_COVERAGE_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["pinning", "retirement_threshold", "error_rate", "workload"],
        summary="Rerun of 2_retirement_threshold at thr=2 only. Newer simulator output: contains [LLC] Baseline Protection Coverage block (needed for pin_off protected_lines_pct).",
    ),
    ExperimentSpec(
        dirname="2_retirement_threshold_only32",
        sheet="2_retirement_threshold_only32",
        config_cols=["pinning", "retirement_threshold", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + PIN_OFF_COVERAGE_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["pinning", "retirement_threshold", "error_rate", "workload"],
        summary="Rerun of 2_retirement_threshold at thr=32 only. Newer output: contains Baseline Protection Coverage block (this is the pin_off thr=32 data fig9 uses for protected lines).",
    ),
    ExperimentSpec(
        dirname="2_retirement_threshold_4_8_16",
        sheet="2_retirement_threshold_4_8_16",
        config_cols=["pinning", "retirement_threshold", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + PIN_OFF_COVERAGE_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["pinning", "retirement_threshold", "error_rate", "workload"],
        summary="Rerun of 2_retirement_threshold at thr={4,8,16}. Newer output: contains Baseline Protection Coverage block.",
    ),
    ExperimentSpec(
        dirname="3_error_way_capacity",
        sheet="3_error_way_capacity",
        config_cols=["allocated_error_ways", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["allocated_error_ways", "error_rate", "workload"],
        summary="Reserved-error-way count sweep at fixed 2 MB LLC, pinning ON. Supplemental sweep; the main fig11 sweep is in 6_llc_way_sweep.",
    ),
    ExperimentSpec(
        dirname="4_llc_size_baseline",
        sheet="4_llc_size_baseline",
        config_cols=["llc_size_mb"],
        metric_cols=BASE_COLS,
        sort_keys=["llc_size_mb", "workload"],
        summary="No-error baseline across LLC sizes (1/2/4/8 MB). Used as the per-workload denominator for fig6/fig7b/fig7c/fig7d/fig9/fig10/fig11.",
    ),
    ExperimentSpec(
        dirname="5_llc_size_sensitivity",
        sheet="5_llc_size_sensitivity",
        config_cols=["llc_size_mb", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["llc_size_mb", "error_rate", "workload"],
        summary="LLC size × CE rate sweep with pinning ON. Drives fig10.",
    ),
    ExperimentSpec(
        dirname="6_llc_way_sweep",
        sheet="6_llc_way_sweep",
        config_cols=["llc_size_mb", "allocated_error_ways", "error_rate"],
        metric_cols=BASE_COLS + ERROR_COLS + PIN_COLS + DERIVED_PROTECTION_COLS,
        sort_keys=["llc_size_mb", "allocated_error_ways", "error_rate", "workload"],
        summary="Reserved-error-way sweep across LLC sizes × CE rates with pinning ON. Drives fig11/fig11b (and the pin_on side of fig9 protected_lines).",
    ),
    ExperimentSpec(
        dirname="7_no_error_way_sweep",
        sheet="7_no_error_way_sweep",
        config_cols=["llc_size_mb", "llc_ways"],
        metric_cols=BASE_COLS,
        sort_keys=["llc_size_mb", "llc_ways", "workload"],
        summary="No-error LLC associativity sweep. Drives fig12.",
    ),
]


# ---------------------------------------------------------------------------
# Glossary — every column that can appear on a sheet, in one place.
# Order is the order they show up on the Definitions sheet.
# ---------------------------------------------------------------------------

GLOSSARY: List[Tuple[str, str, str]] = [
    # (column, units / type, description)

    # --- configuration columns (from filename) ---
    ("pinning",                "on / off",  "LLC Cache Pinning state. 'on' = error lines pinned into reserved high-index ways; 'off' = conventional LLC."),
    ("retirement_threshold",   "int",       "Page-offline retirement threshold (errors-per-page before the page is retired). In 2_retirement_threshold; with pinning=off, thr=2 is the 'Conventional Page Offline' baseline."),
    ("error_rate",             "string",    "Per-access DRAM error injection rate. '1e-5' is benign, '1e-8' is harsh (see Mean-Time-Between-CE column in Definitions)."),
    ("allocated_error_ways",   "int",       "Number of LLC ways reserved for error data (used by 3_error_way_capacity and 6_llc_way_sweep)."),
    ("llc_size_mb",            "MB",        "LLC total capacity for this run."),
    ("llc_ways",               "int",       "LLC associativity (only set in 7_no_error_way_sweep, where 'w<n>' means ways, not error ways)."),

    # --- per-run identity ---
    ("workload",               "string",    "SPEC CPU2017 workload key, e.g. '603.bwaves_s'."),
    ("trace",                  "string",    "Full trace name including the simpoint '-<n>B' suffix."),
    ("filename",               "string",    "Source .txt file under the experiment directory."),
    ("panic",                  "bool",      "True if the simulator printed 'Simulation CPU 0 panic' — IPC may be incomplete/0."),

    # --- ROI metrics ---
    ("ipc",                    "float",     "ROI cumulative IPC of CPU 0 (last 'CPU 0 cumulative IPC:' line)."),
    ("instructions",           "int",       "ROI retired instruction count."),
    ("cycles",                 "int",       "ROI cycle count."),
    ("branch_accuracy_pct",    "%",         "Branch predictor accuracy at end of ROI."),
    ("branch_mpki",            "MPKI",      "Branch mispredictions per 1k retired instructions."),

    # --- LLC counters (CPU 0) ---
    ("llc_load_access",        "int",       "cpu0->LLC LOAD ACCESS counter."),
    ("llc_load_hit",           "int",       "cpu0->LLC LOAD HIT counter."),
    ("llc_load_miss",          "int",       "cpu0->LLC LOAD MISS counter."),
    ("llc_total_access",       "int",       "cpu0->LLC TOTAL ACCESS counter (all access types)."),
    ("llc_total_hit",          "int",       "cpu0->LLC TOTAL HIT counter."),
    ("llc_total_miss",         "int",       "cpu0->LLC TOTAL MISS counter."),
    ("llc_load_mpki",          "MPKI",      "llc_load_miss / instructions * 1000."),
    ("rbmpki",                 "MPKI",      "DRAM ROW_BUFFER_MISS per 1k instructions (sum across channels)."),
    ("page_size_bytes",        "bytes",     "Page size reported by the simulator (4096 or 2097152)."),

    # --- error model summary ---
    ("error_cycle_interval",   "cycles",    "Error injection cadence: simulator advances 'error_cycle_interval' cycles between intervals."),
    ("errors_per_interval",    "int",       "Number of errors fired each interval. (MTBCE ≈ error_cycle_interval / errors_per_interval cycles.)"),
    ("total_error_accesses",   "int",       "Total error-address accesses observed across the ROI."),
    ("total_dram_error_events","int",       "[ERROR] Total DRAM Error Events — every CE the error model fired (cumulative). Includes errors on pages that were later retired (so this can exceed total_known_error_addresses). Used as the numerator-base for fig9's protected_lines_pct."),
    ("total_known_error_addresses", "int",  "Distinct error addresses currently tracked (i.e. not yet retired). Used as 'live' in fig9's protected_lines_pct formula."),
    ("page_retirements_nth_err",     "int", "Pages retired due to the 'Nth error' overflow event in the [ERROR] block (pinning ON)."),
    ("pages_retired",          "int",       "[ERROR] Pages Retired counter — pages retired by the LLC-pinning path (pinning ON)."),
    ("baseline_retirement_threshold","int", "Retirement threshold reported in the no-pinning Baseline block."),
    ("baseline_page_retirements","int",     "Baseline Page Retirements counter — pages retired by the conventional page-offline path (pinning OFF)."),
    ("error_retirement_threshold","int",    "[ERROR] Retirement Threshold (the one actually in force when pinning is ON)."),

    # --- LLC error-way bookkeeping (pinning ON only) ---
    ("allocated_error_ways_per_set","int",  "[LLC] Allocated Error Ways per Set (current allocation in steady state)."),
    ("max_error_ways_per_set", "int",       "[LLC] Max Error Ways per Set (cap, set in config)."),
    ("total_error_way_slots",  "int",       "[LLC] Total Error Way Slots (max_error_ways * num_sets)."),
    ("used_slots",             "int",       "[LLC] Used Slots — error-way slots currently holding an error line. Used as numerator for protected-line coverage."),
    ("used_slots_pct",         "%",         "Used Slots / Total Error Way Slots * 100 (as reported by the simulator)."),
    ("unused_slots",           "int",       "[LLC] Unused Slots."),
    ("unused_slots_pct",       "%",         "Unused Slots / Total Error Way Slots * 100."),
    ("error_way_hits",         "int",       "[LLC] Error Way Hits — runtime hits served from the error way."),
    ("error_way_fills",        "int",       "[LLC] Error Way Fills — error lines installed into the error way."),
    ("error_way_hit_rate_pct", "%",         "[LLC] Error Way Hit Rate — runtime hit rate on error-line accesses (used by fig11b)."),
    ("error_way_evictions",    "int",       "[LLC] Error Way Evictions — error lines evicted from the error way."),
    ("pinned_in_error_way",    "int",       "[LLC] Pinned in Error Way — known error addresses currently resident in the error way (used by fig11 coverage)."),
    ("pinned_in_error_way_pct","%",         "pinned_in_error_way / total_known_error_addresses * 100, as reported by the simulator."),
    ("in_normal_way_unprotected","int",     "[LLC] In Normal Way (unprotected) — known error addresses present in LLC but outside the error way (not protected)."),
    ("not_in_llc_dram_exposed","int",       "[LLC] Not in LLC (DRAM exposed) — known error addresses not resident in LLC (exposed to DRAM CEs)."),

    # --- pinning OFF Baseline Protection Coverage block (subsidiary 2_retirement_threshold_* dirs only) ---
    ("baseline_retired_count_page_offline","int", "[LLC] Retired (page offline) — # of known error addresses on pages that have been retired (page-offline path)."),
    ("baseline_retired_pct_page_offline","%",     "[LLC] Retired (page offline) percent — already-computed retired / (retired + live) * 100. This is the value fig9 plots for the pin-OFF marker."),
    ("baseline_live_still_tracked","int",         "[LLC] Live (still tracked) — # of known error addresses on pages that are NOT yet retired."),

    # --- derived columns (this script computes them) ---
    ("retired_lines_inferred", "int",       "Derived: max(0, total_dram_error_events - total_known_error_addresses). Number of error-address events on pages that have since been retired (those addresses are no longer in the 'known' set). Used as 'retired' in fig9's pin_on formula."),
    ("protected_lines_pct",    "%",         "Derived (fig9): fraction of error addresses currently 'under protection'. "
                                              "Pin ON  → 100 * (pinned_in_error_way + retired_lines_inferred) / (total_known_error_addresses + retired_lines_inferred). "
                                              "Pin OFF → baseline_retired_pct_page_offline (read directly from the simulator's Baseline Protection Coverage block)."),
]

GLOSSARY_KEYS = {row[0] for row in GLOSSARY}


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _last(rx: re.Pattern, text: str):
    """Sim output is duplicated (mid-run + end-of-ROI dumps); always take the last match."""
    matches = list(rx.finditer(text))
    return matches[-1] if matches else None


def parse_result_file(path: str) -> Dict[str, object]:
    """Pull every column listed in GLOSSARY (where present) out of one .txt result."""
    with open(path, "r", errors="replace") as f:
        text = f.read()

    out: Dict[str, object] = {}
    out["panic"] = "Simulation CPU 0 panic" in text

    m = _last(RE_ROI, text)
    if m:
        out["ipc"] = float(m.group(1))
        out["instructions"] = int(m.group(2))
        out["cycles"] = int(m.group(3))

    m = _last(RE_BR, text)
    if m:
        out["branch_accuracy_pct"] = float(m.group(1))
        out["branch_mpki"] = float(m.group(2))

    m = _last(RE_LLC_LOAD, text)
    if m:
        out["llc_load_access"] = int(m.group(1))
        out["llc_load_hit"] = int(m.group(2))
        out["llc_load_miss"] = int(m.group(3))

    m = _last(RE_LLC_TOTAL, text)
    if m:
        out["llc_total_access"] = int(m.group(1))
        out["llc_total_hit"] = int(m.group(2))
        out["llc_total_miss"] = int(m.group(3))

    if out.get("instructions") and out.get("llc_load_miss") is not None:
        out["llc_load_mpki"] = out["llc_load_miss"] * 1000.0 / out["instructions"]

    if out.get("instructions"):
        rb_misses = sum(int(x) for x in RE_ROW_BUFFER_MISS.findall(text))
        if rb_misses > 0:
            out["rbmpki"] = rb_misses * 1000.0 / out["instructions"]

    m = _last(RE_PAGE_SIZE, text)
    if m:
        out["page_size_bytes"] = int(m.group(1))

    # error model
    m = _last(RE_ERR_INTERVAL, text)
    if m:
        out["error_cycle_interval"] = int(m.group(1))
    m = _last(RE_ERR_PER_INT, text)
    if m:
        out["errors_per_interval"] = int(m.group(1))
    m = _last(RE_TOTAL_ERR, text)
    if m:
        out["total_error_accesses"] = int(m.group(1))
    m = _last(RE_TOTAL_DRAM_ERROR_EVENTS, text)
    if m:
        out["total_dram_error_events"] = int(m.group(1))
    m = _last(RE_TOTAL_KNOWN, text)
    if m:
        out["total_known_error_addresses"] = int(m.group(1))

    m = _last(RE_BASELINE_RETIRE, text)
    if m:
        out["baseline_retirement_threshold"] = int(m.group(1))
    m = _last(RE_BASELINE_PAGES, text)
    if m:
        out["baseline_page_retirements"] = int(m.group(1))
    m = _last(RE_ERROR_THRESH, text)
    if m:
        out["error_retirement_threshold"] = int(m.group(1))
    m = _last(RE_ERROR_PAGES_NTH, text)
    if m:
        out["page_retirements_nth_err"] = int(m.group(1))
    m = _last(RE_ERROR_PAGES_RETIRED, text)
    if m:
        out["pages_retired"] = int(m.group(1))

    # error-way bookkeeping
    m = _last(RE_ALLOC, text)
    if m:
        out["allocated_error_ways_per_set"] = int(m.group(1))
    m = _last(RE_MAX, text)
    if m:
        out["max_error_ways_per_set"] = int(m.group(1))
    m = _last(RE_TOTAL_SLOTS, text)
    if m:
        out["total_error_way_slots"] = int(m.group(1))
    m = _last(RE_USED_SLOTS, text)
    if m:
        out["used_slots"] = int(m.group(1))
        out["used_slots_pct"] = float(m.group(2))
    m = _last(RE_UNUSED_SLOTS, text)
    if m:
        out["unused_slots"] = int(m.group(1))
        out["unused_slots_pct"] = float(m.group(2))
    m = _last(RE_EW_HITS, text)
    if m:
        out["error_way_hits"] = int(m.group(1))
    m = _last(RE_EW_FILLS, text)
    if m:
        out["error_way_fills"] = int(m.group(1))
    m = _last(RE_EW_HITRATE, text)
    if m:
        out["error_way_hit_rate_pct"] = float(m.group(1))
    m = _last(RE_EW_EVICT, text)
    if m:
        out["error_way_evictions"] = int(m.group(1))
    m = _last(RE_PINNED, text)
    if m:
        out["pinned_in_error_way"] = int(m.group(1))
        out["pinned_in_error_way_pct"] = float(m.group(2))
    m = _last(RE_IN_NORMAL_WAY, text)
    if m:
        out["in_normal_way_unprotected"] = int(m.group(1))
    m = _last(RE_NOT_IN_LLC, text)
    if m:
        out["not_in_llc_dram_exposed"] = int(m.group(1))

    # Pin OFF Baseline Protection Coverage block (subsidiary 2_retirement_threshold_* dirs)
    m = _last(RE_BASELINE_RETIRED, text)
    if m:
        out["baseline_retired_count_page_offline"] = int(m.group(1))
        out["baseline_retired_pct_page_offline"] = float(m.group(2))
    m = _last(RE_BASELINE_LIVE, text)
    if m:
        out["baseline_live_still_tracked"] = int(m.group(1))

    # --- derived columns (fig9 protected_lines_pct) -------------------------
    pinned = out.get("pinned_in_error_way")
    live = out.get("total_known_error_addresses")
    dram_events = out.get("total_dram_error_events")
    baseline_pct = out.get("baseline_retired_pct_page_offline")

    if dram_events is not None and live is not None:
        retired = max(0, dram_events - live)
        out["retired_lines_inferred"] = retired
        if pinned is not None:
            denom = live + retired
            if denom > 0:
                out["protected_lines_pct"] = 100.0 * (pinned + retired) / denom

    # Pin OFF fallback: take the percent the simulator already computed.
    if "protected_lines_pct" not in out and baseline_pct is not None:
        out["protected_lines_pct"] = baseline_pct

    return out


def _coerce_config_value(key: str, raw: str):
    """Numeric config fields stay numeric; rate-like fields stay as strings for filter clarity."""
    if key == "error_rate":
        return raw
    if key == "pinning":
        return raw
    try:
        return int(raw)
    except ValueError:
        try:
            return float(raw)
        except ValueError:
            return raw


def parse_experiment(exp: ExperimentSpec) -> List[Dict[str, object]]:
    parser = FILENAME_PARSERS[exp.dirname]
    d = os.path.join(RESULTS_ROOT, exp.dirname)
    rows: List[Dict[str, object]] = []
    if not os.path.isdir(d):
        print(f"[skip] missing dir: {d}", file=sys.stderr)
        return rows

    for fn in sorted(os.listdir(d)):
        if not fn.endswith(".txt") or fn == "run_log.txt":
            continue
        m = parser.match(fn)
        if not m:
            print(f"[warn] unrecognized filename: {exp.dirname}/{fn}", file=sys.stderr)
            continue
        parts = m.groupdict()
        trace = parts.pop("trace")
        wm = WORKLOAD_RE.match(trace)
        workload = wm.group(1) if wm else trace

        row: Dict[str, object] = {
            "workload": workload,
            "trace": trace,
            "filename": fn,
        }
        for k, v in parts.items():
            row[k] = _coerce_config_value(k, v)

        row.update(parse_result_file(os.path.join(d, fn)))
        rows.append(row)

    # stable sort by the experiment's sort keys
    def sort_key(r):
        return tuple((r.get(k) is None, r.get(k)) for k in exp.sort_keys)
    rows.sort(key=sort_key)
    return rows


# ---------------------------------------------------------------------------
# XLSX builder
# ---------------------------------------------------------------------------

MTBCE_TABLE = [
    ("1e-5", "36 ms",  "1 CE/hr"),
    ("1e-6", "3.6 ms", "10 CE/hr"),
    ("1e-7", "360 us", "100 CE/hr"),
    ("1e-8", "36 us",  "1000 CE/hr"),
]


def build_xlsx(out_path: str) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
        return 1

    wb = Workbook()
    title_font = Font(bold=True, size=12)
    section_font = Font(bold=True, color="333333")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EDF5")
    wrap = Alignment(vertical="top", wrap_text=True)

    # ---- Definitions sheet -----------------------------------------------
    defs = wb.active
    defs.title = "Definitions"

    defs.append(["normal_evaluation raw data — column glossary"])
    defs["A1"].font = title_font
    defs.append([])

    defs.append(["Experiments (one sheet per row)"])
    defs.cell(row=defs.max_row, column=1).font = section_font
    defs.append(["sheet", "results subdirectory", "what's in it"])
    for cell in defs[defs.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for exp in EXPERIMENTS:
        d = os.path.join(RESULTS_ROOT, exp.dirname)
        if not os.path.isdir(d):
            continue
        n_files = sum(1 for x in os.listdir(d) if x.endswith(".txt") and x != "run_log.txt")
        if n_files == 0:
            continue
        defs.append([exp.sheet, f"results/normal_evaluation/{exp.dirname}/", exp.summary])

    defs.append([])
    defs.append(["Error rate → Mean Time Between CEs"])
    defs.cell(row=defs.max_row, column=1).font = section_font
    defs.append(["error_rate", "MTBCE (trace timeline)", "rate (errors/hour)"])
    for cell in defs[defs.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for rate, mtbce, per_hour in MTBCE_TABLE:
        defs.append([rate, mtbce, per_hour])

    defs.append([])
    defs.append(["Column glossary"])
    defs.cell(row=defs.max_row, column=1).font = section_font
    defs.append(["column", "units / type", "meaning"])
    for cell in defs[defs.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for name, units, desc in GLOSSARY:
        defs.append([name, units, desc])

    defs.column_dimensions["A"].width = 34
    defs.column_dimensions["B"].width = 22
    defs.column_dimensions["C"].width = 110
    for row in defs.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = wrap
    defs.freeze_panes = "A2"

    # ---- One sheet per experiment ----------------------------------------
    for exp in EXPERIMENTS:
        if not os.path.isdir(os.path.join(RESULTS_ROOT, exp.dirname)):
            print(f"[skip] missing dir: results/normal_evaluation/{exp.dirname}", file=sys.stderr)
            continue
        rows = parse_experiment(exp)
        if not rows:
            print(f"[skip] empty dir: results/normal_evaluation/{exp.dirname}", file=sys.stderr)
            continue
        ws = wb.create_sheet(exp.sheet[:31])

        columns = exp.config_cols + exp.metric_cols
        # warn on any unknown column (would mean a glossary gap)
        for col in columns:
            if col not in GLOSSARY_KEYS:
                print(f"[warn] {exp.sheet}: column '{col}' not in Definitions glossary", file=sys.stderr)

        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for r in rows:
            ws.append([r.get(c) for c in columns])

        for col_idx, name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(28, len(name) + 2))

        ws.freeze_panes = "A2"
        last_col_letter = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}{1 + len(rows)}"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    return 0


def main() -> int:
    return build_xlsx(OUT_XLSX)


if __name__ == "__main__":
    sys.exit(main())
